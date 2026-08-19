import tempfile
import unittest
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

from analyzer import analyze_official_docs, analyze_social_comments
from classification_engine import build_classification_outputs
from fund_parser import build_fund_summary, extract_funds, group_underlying_funds
from master_skills import build_master_skill_catalog, build_master_skill_outputs
from market_data import _normalise_nav_rows
from mention_pipeline import (
    build_audit_summary,
    build_content_hash,
    build_mentions_from_social_results,
    build_review_queue,
    redact_pii,
)
from product_master import build_product_master_data
from peer_scoring import build_peer_score_outputs
from report_writer import write_report
from schemas import FundCompanyMention, SourceRef, SourceType
from search_docs import extract_document_date, score_document, score_identity_match
from social_search import build_company_search_units, filter_recent_social, read_social_comments_csv
from structured_products import build_structured_product_outputs
from window import default_window, subtract_months


class FundPipelineTests(unittest.TestCase):
    def test_extracts_and_groups_share_classes(self):
        raw = pd.DataFrame({
            '基金代码': ['968000.OF', '968001.OF', '968002.OF'],
            '基金名称': ['测试亚洲债券人民币份额', '测试亚洲债券美元份额', '测试高股息股票基金'],
            'Product name': [
                'ABC Asian Total Return Bond Fund RMB',
                'ABC Asian Total Return Bond Fund USD',
                'XYZ Asian Dividend Equity Fund Acc',
            ],
            'ISIN': ['LU1234567890', 'LU1234567891', 'LU1234567892'],
        })
        funds = extract_funds(raw)
        share_df, base_df = group_underlying_funds(funds)
        summary = build_fund_summary(share_df, base_df)

        self.assertEqual(len(share_df), 3)
        self.assertEqual(len(base_df), 2)
        self.assertEqual(int(summary.iloc[0]['share_product_count']), 3)
        self.assertIn('测试亚洲债券人民币份额', summary.iloc[0]['fund_names'])
        self.assertIn('未知基金公司', summary.iloc[0]['fund_companies'])

    def test_underlying_identity_prefers_stable_code_over_fuzzy_name(self):
        raw = pd.DataFrame({
            '基金代码': ['968100.OF', '968101.OF', '968102.OF'],
            '晨星代码': ['FS00000001', 'FS00000001', 'FS00000002'],
            'Product name': [
                'Alpha Asian Bond Fund CNY ACC',
                'Alpha Asian Fixed Income Portfolio USD DIST',
                'Alpha Asian Bond Fund USD ACC',
            ],
            'ISIN': ['HK0000000100', 'HK0000000101', 'HK0000000102'],
        })
        shares, base = group_underlying_funds(extract_funds(raw))
        self.assertEqual(len(base), 2)
        self.assertEqual(shares.iloc[0]['base_fund_id'], shares.iloc[1]['base_fund_id'])
        self.assertNotEqual(shares.iloc[0]['base_fund_id'], shares.iloc[2]['base_fund_id'])
        self.assertEqual(shares.iloc[1]['identity_resolution_method'], 'morningstar_code')

    def test_company_search_units_group_by_fund_company(self):
        raw = pd.DataFrame({
            '基金代码': ['968000.OF', '968001.OF', '968040.OF'],
            '基金名称': ['摩根亚洲总收益债券基金 累积', '摩根亚洲总收益债券基金 派息', '惠理价值基金'],
            'Product name': [
                'JPMorgan Asian Total Return Bond Fund RMB ACC',
                'JPMorgan Asian Total Return Bond Fund RMB DIST',
                'Value Partners Classic Fund P-CNY',
            ],
            'ISIN': ['HK0000259686', 'HK0000259694', 'HK0000264959'],
        })
        share_df, _ = group_underlying_funds(extract_funds(raw))
        units = build_company_search_units(share_df)

        self.assertEqual(set(units['fund_company']), {'摩根资产管理', '惠理集团'})
        jpm = units[units['fund_company'].eq('摩根资产管理')].iloc[0]
        self.assertIn('968000.OF', jpm['product_code'])
        self.assertIn('JPMorgan', jpm['company_aliases'])

    def test_document_date_and_scoring_prefers_latest_official_monthly(self):
        doc_date, source = extract_document_date('Fund Monthly Report March 2026', 'https://manager.com/report.pdf')
        self.assertEqual(doc_date, '2026-03-01')
        self.assertEqual(source, 'month_name')

        fresh, relevance = score_document({
            'document_date': doc_date,
            'doc_type_guess': 'monthly_report',
            'source_quality': 'official_or_manager',
        }, today=date(2026, 5, 15))
        self.assertGreaterEqual(fresh, 70)
        self.assertGreaterEqual(relevance, 100)

    def test_document_identity_gate_rejects_same_manager_wrong_product(self):
        fund = pd.Series({
            'base_fund_name': 'JPMorgan Asian Dividend',
            'fund_names_en': 'JPMorgan Asian Dividend - PRC CNY ACC',
            'fund_names_cn': '摩根亚洲股息基金 PRC-CNY 累积',
            'isins': 'HK0000431814',
            'product_codes': '968048.OF',
            'morningstar_codes': 'FS00008Z98',
        })
        correct_score, _ = score_identity_match(
            fund,
            'JPMorgan Asia Equity Dividend (acc) - USD',
            'https://am.jpmorgan.com/products/jpmorgan-asia-equity-dividend',
        )
        wrong_score, _ = score_identity_match(
            fund,
            'JPMorgan Investment Funds - Global Dividend Fund',
            'https://am.jpmorgan.com/global-dividend-fund.pdf',
        )
        code_score, method = score_identity_match(
            fund,
            'Fund factsheet',
            'https://example.com/document/HK0000431814.pdf',
        )
        self.assertGreaterEqual(correct_score, 0.68)
        self.assertLess(wrong_score, 0.68)
        self.assertEqual(code_score, 1.0)
        self.assertTrue(method.startswith('identifier:'))

        amundi = pd.Series({
            'base_fund_name': 'Amundi HK - Growth Fund',
            'fund_names_en': 'Amundi HK-Growth Fund M CNY ACC',
            'fund_names_cn': '东方汇理香港增长基金 M CNY 累积',
            'isins': 'HK001122040',
            'product_codes': '968171.OF',
            'morningstar_codes': 'FSUSA0BBOO',
        })
        same_manager_wrong, _ = score_identity_match(
            amundi,
            'AMUNDI FUNDS MULTI-STRATEGY GROWTH',
            'https://www.amundi.com/globaldistributor/product/view/LU1883335165',
        )
        self.assertLess(same_manager_wrong, 0.68)

    def test_official_analysis_uses_docs_and_peers(self):
        base_df = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'base_fund_name': 'ABC Asian Total Return Bond Fund'},
            {'base_fund_id': 'FUND_0002', 'base_fund_name': 'DEF Asian Bond Fund'},
        ])
        docs_df = pd.DataFrame([
            {
                'base_fund_id': 'FUND_0001',
                'title': 'ABC Asian Total Return Bond Fund Factsheet March 2026',
                'snippet': 'duration credit hedged monthly report',
                'text_excerpt': 'The fund uses duration management and hedging.',
                'is_latest_candidate': True,
                'relevance_score': 110,
                'document_date': '2026-03-01',
                'link': 'https://example.com/factsheet.pdf',
                'doc_type_guess': 'factsheet',
                'source_domain': 'example.com',
            }
        ])
        analysis = analyze_official_docs(base_df, docs_df)
        first = analysis[analysis['base_fund_id'].eq('FUND_0001')].iloc[0]
        self.assertEqual(first['category'], '亚洲债券')
        self.assertIn('同类基金', first['peer_comparison'])
        self.assertIn('久期主动管理', first['rare_or_differentiated_strategies'])

    def test_social_csv_filters_old_comments_and_analyzes_pain_points(self):
        path = '/tmp/fund-social-test.csv'
        pd.DataFrame([
            {'fund_company': '摩根资产管理', 'platform': '小红书', 'user_text': '净值跌了但派息还行', 'publish_time': '2026-03-01'},
            {'fund_company': '摩根资产管理', 'platform': '微博', 'user_text': '老评论', 'publish_time': '2024-01-01'},
        ]).to_csv(path, index=False)
        social = filter_recent_social(read_social_comments_csv(path))
        mentions = build_mentions_from_social_results(social, window=(date(2025, 12, 6), date(2026, 6, 6)))
        analysis = analyze_social_comments(mentions)

        self.assertEqual(len(social), 1)
        self.assertEqual(analysis.iloc[0]['fund_company'], '摩根资产管理')
        self.assertEqual(int(analysis.iloc[0]['valid_comment_count']), 1)
        self.assertIn('亏损/回撤', analysis.iloc[0]['top_pain_aspects'])

    def test_window_uses_natural_months(self):
        self.assertEqual(subtract_months(date(2026, 6, 6), 6), date(2025, 12, 6))
        self.assertEqual(default_window(today=date(2026, 6, 6)), (date(2025, 12, 6), date(2026, 6, 6)))
        self.assertEqual(subtract_months(date(2026, 3, 31), 1), date(2026, 2, 28))

    def test_fund_company_mention_schema_validates_required_and_enum_fields(self):
        with self.assertRaises(ValueError):
            FundCompanyMention(id_hash='', source=SourceRef(platform='微博', source_type=SourceType.SOCIAL))
        with self.assertRaises(ValueError):
            SourceRef(platform='微博', source_type='invalid_source')

    def test_mentions_redact_hash_dedupe_and_flag_review(self):
        redacted, changed = redact_pii('我买了摩根基金，电话13800138000，净值跌了')
        self.assertTrue(changed)
        self.assertNotIn('13800138000', redacted)

        h1 = build_content_hash('小红书', 'https://xhslink.com/a', redacted)
        h2 = build_content_hash('小红书', 'https://xhslink.com/a', redacted)
        self.assertEqual(h1, h2)

        social = pd.DataFrame([
            {
                'fund_company': '摩根资产管理',
                'platform': '小红书',
                'source_type': 'user_exported_comment',
                'link': 'https://xhslink.com/a',
                'user_text': '我买了摩根基金，电话13800138000，净值跌了',
                'publish_time': '2026-03-01',
                'product_code': '968000.OF',
                'base_fund_name': '摩根亚洲总收益债券基金',
            },
            {
                'fund_company': '摩根资产管理',
                'platform': '小红书',
                'source_type': 'user_exported_comment',
                'link': 'https://xhslink.com/a',
                'user_text': '我买了摩根基金，电话13800138000，净值跌了',
                'publish_time': '2026-03-01',
                'product_code': '968000.OF',
                'base_fund_name': '摩根亚洲总收益债券基金',
            },
            {
                'fund_company': '摩根资产管理',
                'platform': '抖音',
                'source_type': 'public_search_result',
                'link': 'https://www.douyin.com/search/a',
                'user_text': '摩根基金怎么样',
                'publish_time': '',
            },
        ])
        mentions = build_mentions_from_social_results(social, window=(date(2025, 12, 6), date(2026, 6, 6)))
        self.assertEqual(len(mentions), 3)
        self.assertEqual(mentions.iloc[0]['duplicate_of'], '')
        self.assertEqual(mentions.iloc[1]['duplicate_of'], mentions.iloc[0]['id_hash'])
        self.assertIn('pii', mentions.iloc[0]['risk_flags'])
        self.assertIn('search_result_only', mentions.iloc[2]['risk_flags'])
        self.assertIn('low_confidence_date', mentions.iloc[2]['risk_flags'])

        review = build_review_queue(mentions)
        self.assertEqual(len(review), 1)
        audit = build_audit_summary(social, mentions, window=(date(2025, 12, 6), date(2026, 6, 6)))
        self.assertIn('deduped_comment_count', set(audit['metric']))

    def test_stale_comments_are_flagged_not_silently_dropped(self):
        social = pd.DataFrame([
            {
                'fund_company': '惠理集团',
                'platform': '微博',
                'source_type': 'social',
                'link': 'https://weibo.com/old',
                'user_text': '以前买过，回撤比较大',
                'publish_time': '2024-01-01',
            },
        ])
        mentions = build_mentions_from_social_results(social, window=(date(2025, 12, 6), date(2026, 6, 6)))
        self.assertEqual(len(mentions), 1)
        self.assertIn('stale', mentions.iloc[0]['risk_flags'])
        analysis = analyze_social_comments(mentions)
        self.assertEqual(float(analysis.iloc[0]['recent_hit_rate']), 0)

    def test_report_contains_audit_and_standardized_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_name = Path(tmpdir, 'report.xlsx').name
            output_path = write_report(
                summary_df=pd.DataFrame([{'a': 1}]),
                share_df=pd.DataFrame([{'a': 1}]),
                base_df=pd.DataFrame([{'a': 1}]),
                docs_df=pd.DataFrame([{'a': 1}]),
                official_analysis_df=pd.DataFrame([{'a': 1}]),
                product_master_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                social_df=pd.DataFrame([{'a': 1}]),
                mention_df=pd.DataFrame([{'id_hash': 'abc'}]),
                social_analysis_df=pd.DataFrame([{'fund_company': '摩根资产管理'}]),
                audit_df=pd.DataFrame([{'metric': 'window', 'value': '2025-12-06 至 2026-06-06'}]),
                review_queue_df=pd.DataFrame([{'id_hash': 'abc'}]),
                skill_catalog_df=pd.DataFrame([{'skill_id': 'fund-mechanism-analyzer'}]),
                mechanism_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                pain_map_df=pd.DataFrame([{'investor_pain': '亏损/回撤'}]),
                market_radar_df=pd.DataFrame([{'next_month_narrative': '降息受益'}]),
                evidence_audit_df=pd.DataFrame([{'audit_decision': '可作为初步定位'}]),
                structured_tranche_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                cashflow_cushion_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                structured_risk_audit_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                classification_fact_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                classification_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                fingerprint_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                peer_edges_df=pd.DataFrame([{'fund_a_id': 'FUND_0001'}]),
                classification_review_df=pd.DataFrame([{'base_fund_id': 'FUND_0001'}]),
                output_name=output_name,
            )
            wb = openpyxl.load_workbook(output_path)
            self.assertIn('基金公司评论标准记录', wb.sheetnames)
            self.assertIn('基金产品主数据', wb.sheetnames)
            self.assertIn('采集审计摘要', wb.sheetnames)
            self.assertIn('人工复核队列', wb.sheetnames)
            self.assertIn('大师Skill配置', wb.sheetnames)
            self.assertIn('产品机制穿透', wb.sheetnames)
            self.assertIn('痛点机制映射', wb.sheetnames)
            self.assertIn('下月卖点雷达', wb.sheetnames)
            self.assertIn('反证审计', wb.sheetnames)
            self.assertIn('结构化分层分析', wb.sheetnames)
            self.assertIn('现金流瀑布与安全垫', wb.sheetnames)
            self.assertIn('结构化风险审计', wb.sheetnames)
            self.assertIn('分类事实包', wb.sheetnames)
            self.assertIn('确定性基金分类', wb.sheetnames)
            self.assertIn('策略指纹', wb.sheetnames)
            self.assertIn('同类基金匹配', wb.sheetnames)
            self.assertIn('分类人工复核队列', wb.sheetnames)

    def test_deterministic_classifier_respects_strategy_negative_rules(self):
        base_df = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'base_fund_name': 'Alpha Equity Long Short Fund', 'fund_names_cn': '', 'fund_names_en': 'Alpha Equity Long Short Fund'},
            {'base_fund_id': 'FUND_0002', 'base_fund_name': 'Beta Bond Futures Fund', 'fund_names_cn': '', 'fund_names_en': 'Beta Bond Futures Fund'},
            {'base_fund_id': 'FUND_0003', 'base_fund_name': 'Gamma Systematic Trend Fund', 'fund_names_cn': '', 'fund_names_en': 'Gamma Systematic Trend Fund'},
            {'base_fund_id': 'FUND_0004', 'base_fund_name': 'Delta Multi Asset Fund', 'fund_names_cn': '', 'fund_names_en': 'Delta Multi Asset Fund'},
        ])
        evidence = {
            'FUND_0001': 'Principal investment strategy: The Fund invests at least 80% in equities and uses long/short positions. Target net exposure: 60%.',
            'FUND_0002': 'Principal investment strategy: The Fund invests at least 80% in bonds. Futures and swaps are used only for hedging purposes.',
            'FUND_0003': 'Principal investment strategy: A systematic trend following strategy across global futures. Derivatives are the core return generation instruments.',
            'FUND_0004': 'Principal investment strategy: Strategic asset allocation with at least 30% in equities and at least 30% in bonds.',
        }
        docs_df = pd.DataFrame([
            {
                'base_fund_id': base_id,
                'title': f'{base_id} factsheet',
                'snippet': '',
                'text_excerpt': text,
                'link': f'https://example.com/{base_id}.pdf',
                'document_date': '2026-06-30',
            }
            for base_id, text in evidence.items()
        ])
        outputs = build_classification_outputs(base_df, docs_df, pd.DataFrame())
        classified = outputs['classification_df'].set_index('base_fund_id')

        self.assertEqual(classified.loc['FUND_0001', 'classification_path'], 'Equity/NontraditionalEquity/LongShortNetLong')
        self.assertIn('market_neutral_not_implied_by_long_short', classified.loc['FUND_0001', 'negative_rule_hits'])
        self.assertTrue(classified.loc['FUND_0002', 'classification_path'].startswith('FixedIncome/'))
        self.assertIn('cta_derivatives_hedging_only', classified.loc['FUND_0002', 'negative_rule_hits'])
        self.assertEqual(classified.loc['FUND_0003', 'classification_path'], 'Alternative/CTA/SystematicTrend')
        self.assertEqual(classified.loc['FUND_0003', 'routing_decision'], 'auto_accept')
        self.assertEqual(classified.loc['FUND_0004', 'classification_path'], 'MultiAsset/StrategicAllocation/CrossAsset')
        self.assertIn('multi_asset_not_alternative_multi_strategy', classified.loc['FUND_0004', 'negative_rule_hits'])

    def test_high_yield_peer_graph_uses_fingerprint_and_hard_blocks(self):
        base_df = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'base_fund_name': 'Alpha Asian High Yield Bond Fund', 'fund_names_cn': '', 'fund_names_en': 'Alpha Asian High Yield Bond Fund'},
            {'base_fund_id': 'FUND_0002', 'base_fund_name': 'Beta Asian High Yield Bond Fund', 'fund_names_cn': '', 'fund_names_en': 'Beta Asian High Yield Bond Fund'},
            {'base_fund_id': 'FUND_0003', 'base_fund_name': 'Gamma Global Equity Fund', 'fund_names_cn': '', 'fund_names_en': 'Gamma Global Equity Fund'},
        ])
        docs_df = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'title': 'Factsheet', 'snippet': '', 'text_excerpt': 'Principal investment strategy: At least 80% in Asian high yield bonds below investment grade. Active management.', 'link': 'https://example.com/a.pdf', 'document_date': '2026-06-30'},
            {'base_fund_id': 'FUND_0002', 'title': 'Factsheet', 'snippet': '', 'text_excerpt': 'Principal investment strategy: At least 80% in Asian high yield bonds below investment grade. Active management.', 'link': 'https://example.com/b.pdf', 'document_date': '2026-06-30'},
            {'base_fund_id': 'FUND_0003', 'title': 'Factsheet', 'snippet': '', 'text_excerpt': 'Principal investment strategy: At least 80% in global equities for capital growth.', 'link': 'https://example.com/c.pdf', 'document_date': '2026-06-30'},
        ])
        outputs = build_classification_outputs(base_df, docs_df, pd.DataFrame())
        classified = outputs['classification_df'].set_index('base_fund_id')
        self.assertEqual(classified.loc['FUND_0001', 'classification_l2'], 'HighYield')

        peers = outputs['peer_edges_df']
        same = peers[(peers['fund_a_id'].eq('FUND_0001')) & (peers['fund_b_id'].eq('FUND_0002'))].iloc[0]
        self.assertEqual(same['peer_tier'], 'near')
        self.assertEqual(same['peer_confidence_gate'], 'provisional_downgrade')
        self.assertGreaterEqual(float(same['peer_coverage']), .70)
        blocked = peers[(peers['fund_a_id'].eq('FUND_0001')) & (peers['fund_b_id'].eq('FUND_0003'))].iloc[0]
        self.assertEqual(blocked['peer_tier'], 'not_peer')
        self.assertEqual(blocked['hard_incompatibility'], 'different_l1_asset_or_return_driver')

    def test_master_skill_outputs_configure_fund_analyst_capabilities(self):
        catalog = build_master_skill_catalog()
        self.assertIn('fund-peer-performance-underwriter', set(catalog['skill_id']))
        self.assertIn('P0', set(catalog['priority']))

        base_df = pd.DataFrame([
            {
                'base_fund_id': 'FUND_0001',
                'base_fund_name': 'ABC Asian Total Return Bond Fund',
                'fund_company': '摩根资产管理',
            },
            {
                'base_fund_id': 'FUND_0002',
                'base_fund_name': 'XYZ Asian Dividend Equity Fund',
                'fund_company': '摩根资产管理',
            },
        ])
        official = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'category': '亚洲债券'},
            {'base_fund_id': 'FUND_0002', 'category': '亚洲股息/高息股票'},
        ])
        docs = pd.DataFrame([
            {'base_fund_id': 'FUND_0001', 'title': 'Factsheet May 2026'},
        ])
        social = pd.DataFrame([
            {
                'fund_company': '摩根资产管理',
                'valid_comment_count': 3,
                'top_pain_aspects': '亏损/回撤(2)；分红误解(1)',
                'top_selling_aspects': '高派息/现金流(2)',
            }
        ])
        outputs = build_master_skill_outputs(base_df, official, docs, social)

        mechanism = outputs['mechanism_df']
        self.assertIn('票息收入', mechanism[mechanism['base_fund_id'].eq('FUND_0001')].iloc[0]['return_sources'])
        self.assertIn('历史净值', mechanism.iloc[0]['missing_evidence'])

        pain_map = outputs['pain_map_df']
        self.assertIn('不得把痛点直接包装成卖点', pain_map.iloc[0]['must_not_claim'])

        radar = outputs['market_radar_df']
        self.assertIn('降息受益', set(radar['next_month_narrative']))
        self.assertIn('高派息/现金流', set(radar['next_month_narrative']))

        audit = outputs['evidence_audit_df']
        self.assertIn('官方/第三方资料', audit[audit['base_fund_id'].eq('FUND_0001')].iloc[0]['supporting_evidence'])
        self.assertIn('不可作为最终推荐', audit.iloc[0]['audit_decision'])

    def test_product_master_data_extracts_official_structured_fields(self):
        base_df = pd.DataFrame([{
            'base_fund_id': 'FUND_0001',
            'base_fund_name': 'ABC Asian Total Return Bond Fund',
            'fund_company': '摩根资产管理',
            'share_count': 2,
            'product_codes': '968000.OF, 968001.OF',
            'isins': 'HK0000000001',
            'morningstar_codes': 'FS00000001',
            'fund_names_cn': 'ABC 亚洲债券基金',
            'fund_names_en': 'ABC Asian Total Return Bond Fund CNY HDG ACC',
        }])
        docs_df = pd.DataFrame([
            {
                'base_fund_id': 'FUND_0001',
                'doc_type_guess': 'factsheet',
                'title': 'ABC Fund Factsheet May 2026',
                'link': 'https://example.com/factsheet.pdf',
                'document_date': '2026-05-01',
                'is_latest_candidate': True,
                'relevance_score': 110,
                'freshness_score': 90,
                'snippet': 'Risk indicator 4 Management fee 1.20%',
                'text_excerpt': '''
Investment objective: The Fund aims to provide income and capital growth from Asian bonds.
Benchmark: J.P. Morgan Asia Credit Index
Base currency: USD
Management fee: 1.20%
Ongoing charges figure: 1.55%
Subscription fee: 3.00%
Redemption fee: 0.50%
SRI: 4 out of 7
Top holdings
Tencent 5.20%
HSBC 4.10%
Asset allocation
Bonds 82.00%
Cash 6.00%
Average credit rating: BBB+
Duration: 4.6 years
Yield to maturity: 5.80%
Distribution yield: 4.50%
Maximum drawdown: -8.20%
Dealing frequency: Daily
Settlement period: T+3
Redemption: Daily redemption with normal settlement.
''',
            },
            {
                'base_fund_id': 'FUND_0001',
                'doc_type_guess': 'kfs',
                'title': 'ABC Product Key Facts',
                'link': 'https://example.com/kfs.pdf',
                'document_date': '2026-05-01',
                'is_latest_candidate': False,
                'relevance_score': 100,
                'freshness_score': 80,
                'snippet': '',
                'text_excerpt': '',
            },
        ])
        official = pd.DataFrame([{'base_fund_id': 'FUND_0001', 'category': '亚洲债券'}])
        master = build_product_master_data(base_df, pd.DataFrame(), docs_df, official)
        row = master.iloc[0]

        self.assertEqual(row['latest_factsheet_link'], 'https://example.com/factsheet.pdf')
        self.assertEqual(row['latest_kid_link'], 'https://example.com/kfs.pdf')
        self.assertEqual(row['management_fee'], '1.20%')
        self.assertEqual(row['ongoing_charges'], '1.55%')
        self.assertEqual(row['risk_level'], '4')
        self.assertIn('Tencent', row['top_holdings'])
        self.assertIn('Bonds', row['asset_allocation'])
        self.assertEqual(row['max_drawdown'], '-8.20%')
        self.assertIn('Daily', row['liquidity_terms'])
        self.assertGreaterEqual(int(row['master_data_quality_score']), 80)

    def test_structured_product_tranche_fields_and_outputs(self):
        base_df = pd.DataFrame([{
            'base_fund_id': 'FUND_0001',
            'base_fund_name': 'ABC Structured Income Fund',
            'fund_company': '测试基金公司',
            'share_count': 1,
            'product_codes': 'P00001',
            'isins': 'HK0000000002',
            'morningstar_codes': '',
            'fund_names_cn': 'ABC 结构化收益基金',
            'fund_names_en': 'ABC Structured Income Fund',
        }])
        docs_df = pd.DataFrame([{
            'base_fund_id': 'FUND_0001',
            'doc_type_guess': 'kfs',
            'title': 'ABC Product Key Facts',
            'link': 'https://example.com/kfs.pdf',
            'document_date': '2026-05-01',
            'is_latest_candidate': True,
            'relevance_score': 100,
            'freshness_score': 80,
            'snippet': '',
            'text_excerpt': '''
本产品为结构化资产管理计划，设置优先级、夹层和劣后级。
优先级预期收益 6%，优先接受收益分配。
夹层预期收益 8%。
劣后级享有剩余收益分配权。
劣后安全垫: 20%
预警线: 90%
止损线: 85%
底层资产为标准化债券，交易日每日开放赎回。
''',
        }])
        master = build_product_master_data(
            base_df,
            pd.DataFrame(),
            docs_df,
            pd.DataFrame([{'base_fund_id': 'FUND_0001', 'category': '策略收益'}]),
        )
        row = master.iloc[0]
        self.assertEqual(row['is_structured_product'], '是')
        self.assertIn('优先', row['tranche_structure'])
        self.assertIn('夹层', row['tranche_structure'])
        self.assertIn('劣后', row['tranche_structure'])
        self.assertEqual(row['waterfall_order'], '优先 > 夹层/平层 > 劣后')
        self.assertEqual(row['loss_absorption_order'], '劣后 > 夹层/平层 > 优先')
        self.assertEqual(row['junior_cushion_ratio'], '20%')
        self.assertEqual(row['warning_line'], '90%')
        self.assertEqual(row['stop_loss_line'], '85%')

        outputs = build_structured_product_outputs(master)
        tranche = outputs['structured_tranche_df'].iloc[0]
        self.assertIn('必须明确投资人认购的是优先', tranche['investor_choice_check'])
        cushion = outputs['cashflow_cushion_df'].iloc[0]
        self.assertIn('劣后安全垫: 20%', cushion['senior_coverage_check'])
        audit = outputs['structured_risk_audit_df'].iloc[0]
        self.assertIn('进一步测算', audit['audit_decision'])

    def test_structured_product_flags_fake_subordination_and_low_liquidity(self):
        base_df = pd.DataFrame([{
            'base_fund_id': 'FUND_0002',
            'base_fund_name': 'DEF Structured Equity Fund',
            'fund_company': '测试基金公司',
            'share_count': 1,
            'product_codes': 'P00002',
            'isins': 'HK0000000003',
            'morningstar_codes': '',
            'fund_names_cn': 'DEF 结构化股权基金',
            'fund_names_en': 'DEF Structured Equity Fund',
        }])
        docs_df = pd.DataFrame([{
            'base_fund_id': 'FUND_0002',
            'doc_type_guess': 'kfs',
            'title': 'DEF Product Key Facts',
            'link': 'https://example.com/def-kfs.pdf',
            'document_date': '2026-05-01',
            'is_latest_candidate': True,
            'relevance_score': 100,
            'freshness_score': 80,
            'snippet': '',
            'text_excerpt': '''
本产品为结构化私募股权基金，包含优先级和劣后级。
优先级目标收益 7%。
劣后级权利义务与优先级没有任何区别。
底层资产为未上市股权，低流动性，退出不确定。
''',
        }])
        master = build_product_master_data(
            base_df,
            pd.DataFrame(),
            docs_df,
            pd.DataFrame([{'base_fund_id': 'FUND_0002', 'category': '主题成长股票'}]),
        )
        flags = master.iloc[0]['structured_product_risk_flags']
        self.assertIn('fake_subordination', flags)
        self.assertIn('low_liquidity_stop_loss', flags)
        self.assertIn('missing_stop_loss_line', flags)
        audit = build_structured_product_outputs(master)['structured_risk_audit_df'].iloc[0]
        self.assertIn('阻断', audit['audit_decision'])

    def test_market_data_uses_cumulative_nav_for_distribution_return_proxy(self):
        nav = _normalise_nav_rows({
            'rows': [
                {'fbrq': '2026-02-02 00:00:00', 'jjjz': '0.95', 'ljjz': '1.10'},
                {'fbrq': '2026-01-02 00:00:00', 'jjjz': '1.00', 'ljjz': '1.00'},
            ],
        })
        self.assertEqual(nav.iloc[-1]['adjusted_nav'], 1.10)
        self.assertGreater(nav.iloc[-1]['adjusted_nav'], nav.iloc[-1]['unit_nav'])

    def test_peer_score_publishes_provisional_not_formal_without_benchmark(self):
        base_df = pd.DataFrame([
            {'base_fund_id': 'F1', 'base_fund_name': 'Value Partners Classic Fund', 'fund_names_cn': '惠理价值基金'},
            {'base_fund_id': 'F2', 'base_fund_name': 'Amundi HK-Growth Fund', 'fund_names_cn': '东方汇理香港增长基金'},
            {'base_fund_id': 'F3', 'base_fund_name': 'JPMorgan SAR Hong Kong PRC', 'fund_names_cn': '摩根香港基金'},
        ])
        dates = pd.date_range('2024-01-31', periods=30, freq='ME')
        series_rows = []
        observation_rows = []
        for index, fund in base_df.iterrows():
            series_id = f'S{index + 1}'
            product_code = f'96810{index}.OF'
            series_rows.append({
                'series_id': series_id,
                'product_code': product_code,
                'base_fund_id': fund['base_fund_id'],
                'base_fund_name': fund['base_fund_name'],
                'fund_company': '测试公司',
                'fund_name_cn': fund['fund_names_cn'],
                'fund_name_en': fund['base_fund_name'],
                'isin': f'HK000000000{index}',
                'reported_currency': 'CNY',
                'is_hedged': False,
                'distribution_type': 'accumulation',
                'history_start': dates.min().date().isoformat(),
                'history_end': dates.max().date().isoformat(),
                'history_months': 29,
                'observation_count': len(dates),
                'return_series_field': 'cumulative_nav',
                'return_reconstruction_method': 'vendor_cumulative_nav_proxy',
                'source_name': 'test',
                'source_url': 'https://example.com',
                'source_quality': 'test',
                'series_status': 'usable',
            })
            for offset, nav_date in enumerate(dates):
                nav = 10 * (1 + 0.004 * (index + 1)) ** offset * (1 + 0.01 * ((offset + index) % 3 - 1))
                observation_rows.append({
                    'series_id': series_id,
                    'product_code': product_code,
                    'base_fund_id': fund['base_fund_id'],
                    'date': nav_date,
                    'unit_nav': nav,
                    'cumulative_nav': nav,
                    'adjusted_nav': nav,
                    'reported_currency': 'CNY',
                    'source_name': 'test',
                    'source_url': 'https://example.com',
                    'retrieved_at': '2026-08-19T00:00:00Z',
                })
        classification = pd.DataFrame([
            {'base_fund_id': fund_id, 'classification_path': 'Equity/LongOnly/BroadMarket'}
            for fund_id in base_df['base_fund_id']
        ])
        fingerprint = pd.DataFrame([
            {
                'base_fund_id': fund_id,
                'return_drivers': '["equity_beta"]',
                'primary_assets': '["equity"]',
                'strategy_mechanics': '["long_only"]',
                'geography': '["HongKong"]',
                'sector_theme': '[]',
                'equity_style': '["Value"]' if fund_id == 'F1' else '["Growth"]',
                'credit_quality': '[]',
                'duration_band': '',
                'derivative_role': '',
                'leverage_role': '',
            }
            for fund_id in base_df['base_fund_id']
        ])
        outputs = build_peer_score_outputs(
            base_df,
            pd.DataFrame(series_rows),
            pd.DataFrame(observation_rows),
            classification,
            fingerprint,
            pd.DataFrame({'base_fund_id': base_df['base_fund_id']}),
        )
        scores = outputs['investment_scores_df']
        self.assertEqual(set(scores['score_status']), {'provisional'})
        self.assertTrue(scores['investment_quality_score'].notna().all())
        self.assertFalse(scores['benchmark_available'].any())
        self.assertTrue(scores['blocked_or_provisional_reason'].str.contains('基准').all())

    def test_peer_score_blocks_short_history_instead_of_zero_score(self):
        base_df = pd.DataFrame([{
            'base_fund_id': 'F1',
            'base_fund_name': 'JPMorgan Asia Equity High Income Fund',
            'fund_names_cn': '摩根亚洲高息股票基金',
        }])
        dates = pd.date_range('2026-04-30', periods=5, freq='ME')
        series_master = pd.DataFrame([{
            'series_id': 'S1', 'product_code': '968213.OF', 'base_fund_id': 'F1',
            'base_fund_name': base_df.iloc[0]['base_fund_name'], 'fund_company': '摩根资产管理',
            'fund_name_cn': base_df.iloc[0]['fund_names_cn'], 'fund_name_en': '', 'isin': '',
            'reported_currency': 'CNY', 'is_hedged': True, 'distribution_type': 'accumulation',
            'history_start': dates.min().date().isoformat(), 'history_end': dates.max().date().isoformat(),
            'history_months': 4, 'observation_count': 5, 'return_series_field': 'cumulative_nav',
            'return_reconstruction_method': 'vendor_cumulative_nav_proxy', 'source_name': 'test',
            'source_url': 'https://example.com', 'source_quality': 'test', 'series_status': 'short_history',
        }])
        observations = pd.DataFrame([{
            'series_id': 'S1', 'product_code': '968213.OF', 'base_fund_id': 'F1', 'date': nav_date,
            'unit_nav': 10 + index, 'cumulative_nav': 10 + index, 'adjusted_nav': 10 + index,
            'reported_currency': 'CNY', 'source_name': 'test', 'source_url': 'https://example.com',
            'retrieved_at': '2026-08-19T00:00:00Z',
        } for index, nav_date in enumerate(dates)])
        outputs = build_peer_score_outputs(
            base_df,
            series_master,
            observations,
            pd.DataFrame([{'base_fund_id': 'F1', 'classification_path': 'Equity/LongOnly/Income'}]),
            pd.DataFrame([{'base_fund_id': 'F1'}]),
            pd.DataFrame([{'base_fund_id': 'F1'}]),
        )
        row = outputs['investment_scores_df'].iloc[0]
        self.assertEqual(row['score_status'], 'blocked')
        self.assertTrue(pd.isna(row['investment_quality_score']))
        self.assertIn('少于12个月', row['blocked_or_provisional_reason'])


if __name__ == '__main__':
    unittest.main()
