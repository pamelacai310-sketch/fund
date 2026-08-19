from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR


PERCENT_HEADERS = {
    'annualized_return', 'annualized_volatility', 'downside_volatility',
    'max_drawdown', 'worst_month', 'positive_rolling_12m_rate',
    'return_12m', 'return_36m', 'return_60m', 'risk_free_rate_assumption',
}
SCORE_HEADERS = {
    'investment_quality_score', 'return_score', 'risk_score',
    'risk_adjusted_score', 'consistency_score', 'broad_diagnostic_score',
    'differentiation_score', 'raw_differentiation_score',
    'structural_rarity_score', 'documented_strategy_score',
    'behavior_divergence_score', 'share_experience_score',
}


def _style_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(name='PingFang SC', size=10, color='FFFFFF', bold=True)
    body_font = Font(name='PingFang SC', size=10, color='000000')
    formal_fill = PatternFill('solid', fgColor='E2F0D9')
    provisional_fill = PatternFill('solid', fgColor='FFF2CC')
    blocked_fill = PatternFill('solid', fgColor='FCE4D6')
    for sheet in workbook.worksheets:
        sheet.freeze_panes = 'A2'
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 30
        headers = {}
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            headers[str(cell.value or '')] = cell.column
        sample_end = min(sheet.max_row, 250)
        for column_index in range(1, sheet.max_column + 1):
            header = str(sheet.cell(1, column_index).value or '')
            max_length = len(header)
            for row_index in range(2, sample_end + 1):
                value = sheet.cell(row_index, column_index).value
                if value is not None:
                    max_length = max(max_length, min(80, len(str(value))))
            sheet.column_dimensions[get_column_letter(column_index)].width = min(42, max(10, max_length + 2))
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_index, column_index)
                cell.font = body_font
                cell.alignment = Alignment(vertical='top', wrap_text=max_length > 24)
                if header in PERCENT_HEADERS and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
                elif header in SCORE_HEADERS and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0'
                elif header.endswith('_confidence') or header in {'metric_completeness', 'evidence_completeness', 'component_coverage'}:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'
                if header in {'source_url', 'link', 'latest_document_link', 'evidence_links'} and isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.hyperlink = cell.value
                    cell.style = 'Hyperlink'
        for status_header in ['score_status', 'status', 'coverage_status']:
            status_col = headers.get(status_header)
            if not status_col:
                continue
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_index, status_col)
                value = str(cell.value or '').lower()
                if value == 'formal':
                    cell.fill = formal_fill
                elif value in {'provisional', '12m', '36m', '60m'}:
                    cell.fill = provisional_fill
                elif value in {'blocked', 'insufficient', 'no_data'}:
                    cell.fill = blocked_fill


def write_report(
    summary_df: pd.DataFrame,
    share_df: pd.DataFrame,
    base_df: pd.DataFrame,
    docs_df: pd.DataFrame,
    official_analysis_df: pd.DataFrame,
    product_master_df: pd.DataFrame,
    social_df: pd.DataFrame,
    mention_df: pd.DataFrame,
    social_analysis_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    review_queue_df: pd.DataFrame,
    skill_catalog_df: pd.DataFrame | None = None,
    mechanism_df: pd.DataFrame | None = None,
    pain_map_df: pd.DataFrame | None = None,
    market_radar_df: pd.DataFrame | None = None,
    evidence_audit_df: pd.DataFrame | None = None,
    structured_tranche_df: pd.DataFrame | None = None,
    cashflow_cushion_df: pd.DataFrame | None = None,
    structured_risk_audit_df: pd.DataFrame | None = None,
    classification_fact_df: pd.DataFrame | None = None,
    classification_df: pd.DataFrame | None = None,
    fingerprint_df: pd.DataFrame | None = None,
    peer_edges_df: pd.DataFrame | None = None,
    classification_review_df: pd.DataFrame | None = None,
    series_master_df: pd.DataFrame | None = None,
    observations_df: pd.DataFrame | None = None,
    market_provenance_df: pd.DataFrame | None = None,
    source_coverage_df: pd.DataFrame | None = None,
    market_review_df: pd.DataFrame | None = None,
    market_audit_df: pd.DataFrame | None = None,
    series_metrics_df: pd.DataFrame | None = None,
    underlying_metrics_df: pd.DataFrame | None = None,
    investment_scores_df: pd.DataFrame | None = None,
    differentiation_scores_df: pd.DataFrame | None = None,
    share_experience_df: pd.DataFrame | None = None,
    peer_ranking_df: pd.DataFrame | None = None,
    peer_pool_definition_df: pd.DataFrame | None = None,
    scoring_blocked_df: pd.DataFrame | None = None,
    scoring_methodology_df: pd.DataFrame | None = None,
    output_name: str = 'fund_research_report.xlsx',
) -> Path:
    output_path = OUTPUT_DIR / output_name
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='基金清单摘要', index=False)
        share_df.to_excel(writer, sheet_name='全部份额', index=False)
        base_df.to_excel(writer, sheet_name='底层基金统计', index=False)
        docs_df.to_excel(writer, sheet_name='官方资料搜索结果', index=False)
        official_analysis_df.to_excel(writer, sheet_name='官方资料分析', index=False)
        product_master_df.to_excel(writer, sheet_name='基金产品主数据', index=False)
        social_df.to_excel(writer, sheet_name='基金公司社媒搜索结果', index=False)
        mention_df.to_excel(writer, sheet_name='基金公司评论标准记录', index=False)
        social_analysis_df.to_excel(writer, sheet_name='基金公司痛点卖点', index=False)
        audit_df.to_excel(writer, sheet_name='采集审计摘要', index=False)
        review_queue_df.to_excel(writer, sheet_name='人工复核队列', index=False)
        if skill_catalog_df is not None:
            skill_catalog_df.to_excel(writer, sheet_name='大师Skill配置', index=False)
        if mechanism_df is not None:
            mechanism_df.to_excel(writer, sheet_name='产品机制穿透', index=False)
        if pain_map_df is not None:
            pain_map_df.to_excel(writer, sheet_name='痛点机制映射', index=False)
        if market_radar_df is not None:
            market_radar_df.to_excel(writer, sheet_name='下月卖点雷达', index=False)
        if evidence_audit_df is not None:
            evidence_audit_df.to_excel(writer, sheet_name='反证审计', index=False)
        if structured_tranche_df is not None:
            structured_tranche_df.to_excel(writer, sheet_name='结构化分层分析', index=False)
        if cashflow_cushion_df is not None:
            cashflow_cushion_df.to_excel(writer, sheet_name='现金流瀑布与安全垫', index=False)
        if structured_risk_audit_df is not None:
            structured_risk_audit_df.to_excel(writer, sheet_name='结构化风险审计', index=False)
        if classification_fact_df is not None:
            classification_fact_df.to_excel(writer, sheet_name='分类事实包', index=False)
        if classification_df is not None:
            classification_df.to_excel(writer, sheet_name='确定性基金分类', index=False)
        if fingerprint_df is not None:
            fingerprint_df.to_excel(writer, sheet_name='策略指纹', index=False)
        if peer_edges_df is not None:
            peer_edges_df.to_excel(writer, sheet_name='同类基金匹配', index=False)
        if classification_review_df is not None:
            classification_review_df.to_excel(writer, sheet_name='分类人工复核队列', index=False)
        if series_master_df is not None:
            series_master_df.to_excel(writer, sheet_name='行情序列主表', index=False)
        if observations_df is not None:
            observations_df.to_excel(writer, sheet_name='净值与总回报序列', index=False)
        if market_provenance_df is not None:
            market_provenance_df.to_excel(writer, sheet_name='数据来源与证据', index=False)
        if source_coverage_df is not None:
            source_coverage_df.to_excel(writer, sheet_name='数据覆盖率', index=False)
        if market_review_df is not None:
            market_review_df.to_excel(writer, sheet_name='行情人工复核', index=False)
        if market_audit_df is not None:
            market_audit_df.to_excel(writer, sheet_name='行情采集审计', index=False)
        if series_metrics_df is not None:
            series_metrics_df.to_excel(writer, sheet_name='全部份额收益风险指标', index=False)
        if underlying_metrics_df is not None:
            underlying_metrics_df.to_excel(writer, sheet_name='底层基金收益风险指标', index=False)
        if investment_scores_df is not None:
            investment_scores_df.to_excel(writer, sheet_name='底层基金综合评分', index=False)
        if differentiation_scores_df is not None:
            differentiation_scores_df.to_excel(writer, sheet_name='差异化特征评分', index=False)
        if share_experience_df is not None:
            share_experience_df.to_excel(writer, sheet_name='份额人民币体验', index=False)
        if peer_ranking_df is not None:
            peer_ranking_df.to_excel(writer, sheet_name='同类基金排名', index=False)
        if peer_pool_definition_df is not None:
            peer_pool_definition_df.to_excel(writer, sheet_name='同类池定义', index=False)
        if scoring_blocked_df is not None:
            scoring_blocked_df.to_excel(writer, sheet_name='评分受阻原因', index=False)
        if scoring_methodology_df is not None:
            scoring_methodology_df.to_excel(writer, sheet_name='评分方法说明', index=False)
        _style_workbook(writer)
    return output_path
